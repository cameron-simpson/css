#!/usr/bin/python -tt
#
# Utility functions for CSV files, particularly to provide consistent
# decoding in python 2 and 3.
#       - Cameron Simpson <cs@cskk.id.au> 02may2013
#

''' Utility functions for CSV files.

    In python 2 the stdlib CSV reader reads 8 bit byte data and returns str objects;
    these need to be decoded into unicode objects.
    In python 3 the stdlib CSV reader reads an open text file and returns str
    objects (== unicode).
    So we provide `csv_reader()` generators to yield rows containing unicode.
'''

from __future__ import absolute_import, print_function
import csv
import sys
from cs.deco import strable
from cs.logutils import warning
from cs.mappings import named_column_tuples
from cs.pfx import Pfx

__version__ = '20220606-post'

DISTINFO = {
    'description':
    "CSV file related facilities",
    'keywords': ["python2", "python3"],
    'classifiers': [
        "Programming Language :: Python",
        "Programming Language :: Python :: 2",
        "Programming Language :: Python :: 3",
    ],
    'install_requires': ['cs.deco', 'cs.logutils', 'cs.mappings', 'cs.pfx'],
}

if sys.hexversion >= 0x03000000:
  # python 3 onwards

  @strable
  def csv_reader(fp, encoding='utf-8', errors='replace', **kw):
    ''' Read the file `fp` using csv.reader.
        `fp` may also be a filename.
        Yield the rows.

        Warning: _ignores_ the `encoding` and `errors` parameters
        because `fp` should already be decoded.
    '''
    for row in csv.reader(fp, **kw):
      yield row

  def csv_writerow(csvw, row, encoding='utf-8'):
    ''' Write the supplied row as strings encoded with the supplied `encoding`,
        default 'utf-8'.
    '''
    with Pfx("csv_writerow(csvw=%s, row=%r, encoding=%r)", csvw, row,
             encoding):
      return csvw.writerow(row)

else:
  # python 2 compatability code

  @strable
  def csv_reader(fp, encoding='utf-8', errors='replace', **kw):
    ''' Read the file `fp` using csv.reader and decode the str
        fields into unicode using the supplied `encoding`,
        default "utf-8".
        `fp` may also be a filename.
        Yield the rows after decoding.
    '''
    r = csv.reader(fp, **kw)
    for row in r:
      for i, value in enumerate(row):
        if isinstance(value, str):
          # transmute str (== bytes) to unicode
          try:
            value = value.decode(encoding)
          except UnicodeDecodeError as e:
            warning("%s, using errors=%s", e, errors)
            value = value.decode('utf-8', errors=errors)
          row[i] = value
      yield row

  def csv_writerow(csvw, row, encoding='utf-8'):
    ''' Write the supplied row as strings encoded with the supplied `encoding`,
        default 'utf-8'.
    '''
    csvw.writerow([unicode(value).encode(encoding) for value in row])

def csv_import(
    fp,
    class_name=None,
    column_names=None,
    computed=None,
    preprocess=None,
    mixin=None,
    snake_case=False,
    **kw
):
  ''' Read CSV data where the first row contains column headers.
      Returns a row namedtuple factory and an iterable of instances.

      Parameters:
      * `fp`: a file object containing CSV data, or the name of such a file
      * `class_name`: optional class name for the namedtuple subclass
        used for the row data.
      * `column_names`: optional iterable of column headings; if
        provided then the file is not expected to have internal column
        headings
      * `computed`: optional keyword parameter providing a mapping
        of str to functions of `self`; these strings are available
        via __getitem__
      * `preprocess`: optional keyword parameter providing a callable
        to modify CSV rows before they are converted into the namedtuple.
        It receives a context object and the data row. It may return
        the row (possibly modified), or None to drop the row.
      * `mixin`: an optional mixin class for the generated namedtuple subclass
        to provide extra methods or properties

      All other keyword parameters are passed to csv_reader(). This
      is a very thin shim around `cs.mappings.named_column_tuples`.

      Examples:

            >>> rowtype, rows = csv_import(['a, b', '1,2', '3,4'], class_name='Example_AB')
            >>> rowtype     #doctest: +ELLIPSIS
            <function named_row_tuple.<locals>.factory at ...>
            >>> list(rows)
            [Example_AB(a='1', b='2'), Example_AB(a='3', b='4')]

            >>> rowtype, rows = csv_import(['1,2', '3,4'], class_name='Example_DEFG', column_names=['D E', 'F G '])
            >>> list(rows)
            [Example_DEFG(d_e='1', f_g='2'), Example_DEFG(d_e='3', f_g='4')]
  '''
  return named_column_tuples(
      csv_reader(fp, **kw),
      class_name=class_name,
      column_names=column_names,
      computed=computed,
      preprocess=preprocess,
      mixin=mixin,
      snake_case=snake_case,
  )

def xl_import(workbook, sheet_name=None, skip_rows=0, **kw):
  ''' Read the named `sheet_name` from the Excel XLSX file named
      `filename` as for `csv_import`.
      Returns a row namedtuple factory and an iterable of instances.

      Parameters:
      * `workbook`: Excel work book from which to load the sheet; if
        this is a str then the work book is obtained from
        openpyxl.load_workbook()
      * `sheet_name`: optional name of the work book sheet
        whose data should be imported;
        the default (`None`) selects the active worksheet

      Other keyword parameters are as for cs.mappings.named_column_tuples.

      NOTE: this function requires the `openpyxl` module to be available.
  '''
  if isinstance(workbook, str):
    from openpyxl import load_workbook
    wb_filename = workbook
    with Pfx(wb_filename):
      workbook = load_workbook(filename=wb_filename, read_only=True)
      return xl_import(workbook, sheet_name, skip_rows=skip_rows, **kw)
  if sheet_name is None:
    worksheet = workbook.active
    if worksheet is None:
      worksheet = workbook[workbook.get_sheet_names()[0]]
  else:
    worksheet = workbook[sheet_name]
  return named_column_tuples(
      (
          [cell.value
           for cell in row]
          for ri, row in enumerate(worksheet)
          if ri >= skip_rows
      ), **kw
  )

if __name__ == '__main__':
  args = sys.argv[1:]
  if not args:
    raise ValueError("missing filename")
  for filename in args:
    print(filename)
    with Pfx(filename):
      if filename.endswith('.csv'):
        with open(filename, 'r') as csvfp:
          cls, rows = csv_import(csvfp)
          for rownum, row in enumerate(rows, 1):
            print(filename, rownum, row)
      elif filename.endswith('.xlsx'):
        from openpyxl import load_workbook
        workbook = load_workbook(filename=filename, read_only=True)
        for wb_sheet_name in workbook.get_sheet_names():
          with Pfx(wb_sheet_name):
            # presume row 1 in some kind of title and column names are row 2
            cls, rows = xl_import(workbook, wb_sheet_name, skip_rows=1)
            for rownum, row in enumerate(rows, 1):
              print(filename, wb_sheet_name, rownum, row)
      else:
        raise ValueError('not a .csv or .xlsx file')
    print()
